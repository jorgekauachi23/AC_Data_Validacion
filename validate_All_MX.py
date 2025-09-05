
import pandas as pd
import os
import argparse
import tempfile
import shutil
from openpyxl.styles import Font

# === CLI Arguments ===
parser = argparse.ArgumentParser(description="Validación de Ventas MX")
group = parser.add_mutually_exclusive_group()
group.add_argument('--completo', action='store_true', help='Procesar Part1 a Part4 (default)')
group.add_argument('--precierre', action='store_true', help='Procesar solo Part1 a Part3')
args = parser.parse_args()

# === Setup ===
base_input = "Input"
num_parts = 3 if args.precierre else 4
part_dirs = [os.path.join(base_input, f"Part{i}") for i in range(1, num_parts + 1)]

output_dir = "Output"
os.makedirs(output_dir, exist_ok=True)

file_map = {
    'Volume': ('volume_sales', 13),
    'Revenue': ('revenue_sales', 13),
    'Transactions': ('stddisc_transaction', 14)
}

CLIENTE_VENDING = 231013

def load_csv_chunked(filepath, col_idx, cliente_id):
    total_rows = []
    vending_rows = []
    vending_by_mat_rows = []

    chunks = pd.read_csv(filepath, header=None, chunksize=500_000)

    for chunk in chunks:
        chunk = chunk.dropna(subset=[0, 2, 3, col_idx])
        chunk[col_idx] = pd.to_numeric(chunk[col_idx], errors='coerce')

        total_rows.append(chunk[[0, col_idx]])

        vending_filtered = chunk[chunk[2] == cliente_id]
        if not vending_filtered.empty:
            vending_rows.append(vending_filtered[[0, col_idx]])
            vending_by_mat_rows.append(vending_filtered[[0, 3, col_idx]])

    total_df = pd.concat(total_rows).groupby(0)[col_idx].sum().reset_index() if total_rows else pd.DataFrame()
    vending_total = pd.concat(vending_rows).groupby(0)[col_idx].sum().reset_index() if vending_rows else pd.DataFrame()
    vending_by_mat = pd.concat(vending_by_mat_rows).groupby([0, 3])[col_idx].sum().reset_index()

    return total_df, vending_total, vending_by_mat

def extract_data(metric_key, pattern, col_idx):
    total_df = pd.DataFrame()
    vending_total = pd.DataFrame()
    vending_by_mat = pd.DataFrame()

    for part_dir in part_dirs:
        if not os.path.isdir(part_dir):
            print(f"⚠️ Folder {part_dir} not found. Skipping.")
            continue

        matched_file = next((f for f in os.listdir(part_dir) if pattern in f.lower()), None)
        if not matched_file:
            print(f"⚠️ No {pattern} file found in {part_dir}.")
            continue

        print(f"📄 Processing {matched_file} from {part_dir} for {metric_key}")
        filepath = os.path.join(part_dir, matched_file)

        t_df, v_df, v_mat_df = load_csv_chunked(filepath, col_idx, CLIENTE_VENDING)
        total_df = pd.concat([total_df, t_df], ignore_index=True)
        vending_total = pd.concat([vending_total, v_df], ignore_index=True)
        vending_by_mat = pd.concat([vending_by_mat, v_mat_df], ignore_index=True)

    total_df = total_df.groupby(0).sum().reset_index() if not total_df.empty else total_df
    vending_total = vending_total.groupby(0).sum().reset_index() if not vending_total.empty else vending_total
    if metric_key != 'Volume' or vending_by_mat.empty:
        vending_by_mat = pd.DataFrame()


    print(f"   ✅ Rows total_df: {len(total_df)} | vending_total: {len(vending_total)} | vending_by_mat: {len(vending_by_mat)}")
    return total_df, vending_total, vending_by_mat

def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# === Excel Export ===
mode = "precierre" if args.precierre else "completo"
excel_filename = f"validation_MX_{mode}.xlsx"
final_excel_path = os.path.join(output_dir, excel_filename)

if os.path.exists(final_excel_path):
    try:
        os.remove(final_excel_path)
        print(f"🧹 Removed existing {final_excel_path}")
    except Exception as e:
        print(f"⚠️ Could not remove old file ({e}). Trying temp write anyway.")

tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
os.close(tmp_fd)

print(f"📝 Writing workbook to temp path: {tmp_path}")

with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
    for metric_name, (pattern, col_idx) in file_map.items():
        total_df, vending_total_df, vending_by_mat_df = extract_data(metric_name, pattern, col_idx)

        if total_df.empty:
            print(f"❌ No data for {metric_name} -- sheet skipped.")
            continue

        total_df = total_df.copy()
        total_df.columns = ["Date", metric_name]
        total_df.to_excel(writer, sheet_name=metric_name, index=False, startrow=1, startcol=0)
        ws = writer.sheets[metric_name]
        ws.cell(row=1, column=1, value="Venta Total").font = Font(bold=True)

        if not vending_total_df.empty:
            vending_total_df = vending_total_df.copy()
            vending_total_df.columns = ["Date", metric_name]
            vending_total_df.to_excel(writer, sheet_name=metric_name, index=False, startrow=1, startcol=5)
            ws.cell(row=1, column=6, value="Filtro 231013").font = Font(bold=True)

        if metric_name == 'Volume' and not vending_by_mat_df.empty:
            vending_by_mat_df = vending_by_mat_df.copy()
            vending_by_mat_df.columns = ["Date", "SKU", metric_name]
            vending_by_mat_df.to_excel(writer, sheet_name=metric_name, index=False, startrow=1, startcol=10)
            ws.cell(row=1, column=11, value="Filtro 231013 por SKU").font = Font(bold=True)

        auto_fit_columns(ws)

try:
    shutil.move(tmp_path, final_excel_path)
    print(f"✅ Excel summary saved to: {final_excel_path}")
except Exception as e:
    print(f"❌ Failed to move workbook into Output: {e}")
    print(f"   Temp file still at: {tmp_path}")
